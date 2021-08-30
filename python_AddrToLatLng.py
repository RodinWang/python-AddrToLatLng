import argparse
import urllib
import requests
import json
from openpyxl import load_workbook

"""
get_latitude_longtitude(api_key, address)
get Latitude and Longitude with google API

api_key - Google Map API key
address - Address needs to be converted

"""
def get_latitude_longtitude(api_key, address):

    # Combine URL and API Key
    address = urllib.parse.quote(address)
    url = "https://maps.googleapis.com/maps/api/geocode/json?address=" + address + '&key=' + api_key

    lat, lng = 0, 0
    try:
        # request google API
        response = requests.get(url)
        if response.status_code == 200:

            # Get Data from response
            js = json.loads(response.text)
            location = js["results"][0]["geometry"]["location"]
            lat = location["lat"]
            lng = location["lng"]
    except:
        print("Get Geometry Error!")

    return lat, lng

"""
    Main Fucntion
    Read the address from the input file and get the latitude & longitude
    Then save to a new file.
"""
def main(api_key, input_file, output_file, workbook, addr_column, lat_column, lng_column):
    # Load file and get sheets
    wb = load_workbook(input_file)
    sheet = wb[workbook]
    print("Sheet has %d row" % (sheet.max_row))

    # start converting 
    for i in range(2, sheet.max_row + 1):
        print("Row = " + str(i))

        # Get Address
        addr = sheet.cell(row=i, column=addr_column).value
        print("Address = ", addr)

        # Get Latitude & Longtitude
        Latitude, Longitude = get_latitude_longtitude(api_key, addr)
        print("Latitude = ", Latitude)
        print("Longitude = ", Longitude)

        # Save Latitude & Longtitude Value
        excel_Latitude = sheet.cell(row=i, column=lat_column)
        excel_Longitude = sheet.cell(row=i, column=lng_column)
        excel_Latitude.value = Latitude
        excel_Longitude.value = Longitude

    # Save file
    wb.save(output_file)

if __name__ == '__main__':

    # argument parsing
    parser = argparse.ArgumentParser(description='Convert your address to Geometry using Google Geocoding API.')
    parser.add_argument("-k", "--api-key", dest="api_key", required=True, help="Input your google api key")
    parser.add_argument("-i", "--input-file", dest="input_file", default="./data/input.xlsx", help="Input xlsx file(e.g. \"./data/input.xlsx\")")
    parser.add_argument("-o", "--output-file", dest="output_file", default="./outputs/result.xlsx", help="output xlsx file(e.g. \"./outputs/result.xlsx\")")
    parser.add_argument("-wb", "--workbook", dest="workbook", default='工作表1', help="Select a workbook in xlsx file. Default='工作表1'")
    parser.add_argument("--addr_col", dest="addr_column", type=int, default=4, help="Address column in xlsx file. Default=4")
    parser.add_argument("--lat_col", dest="lat_column", type=int, default=11, help="Latitude column going to add in xlsx file. Default=11")
    parser.add_argument("--lng_col", dest="lng_column", type=int, default=12, help="Longitude column going to add in xlsx file. Default=12")

    # Get argument
    args = parser.parse_args()
    api_key = args.api_key
    input_file = args.input_file
    output_file = args.output_file
    workbook = args.workbook
    addr_column = args.addr_column
    lat_column = args.lat_column
    lng_column = args.lng_column

    # main process
    main(api_key, input_file, output_file, workbook, addr_column, lat_column, lng_column)

